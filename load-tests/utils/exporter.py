from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from pathlib import Path
from datetime import datetime
from statistics import mean, median, stdev
from typing import List, Dict

def export_results_to_excel(results: List[Dict], filename_prefix="locust_results", folder="load-tests"):
    if not results:
        print("⚠️ No results to export.")
        return

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_detail = wb.create_sheet("All Results")
    ws_status = wb.create_sheet("Status Counts")
    ws_perf = wb.create_sheet("Performance Dist.")

    # All Results
    headers = list(results[0].keys())
    ws_detail.append(headers)
    for row in results:
        ws_detail.append([row[key] for key in headers])

    # Status Counts
    status_codes = {}
    for row in results:
        code = row["status_code"]
        status_codes[code] = status_codes.get(code, 0) + 1

    ws_status.append(["Status Code", "Count"])
    for code, count in sorted(status_codes.items()):
        ws_status.append([str(code), count])

    chart1 = BarChart()
    chart1.title = "HTTP Status Breakdown"
    chart1.y_axis.title = "Count"
    chart1.x_axis.title = "Status Code"
    data = Reference(ws_status, min_col=2, min_row=1, max_row=1 + len(status_codes))
    cats = Reference(ws_status, min_col=1, min_row=2, max_row=1 + len(status_codes))
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    ws_status.add_chart(chart1, "D2")

    # Summary
    response_times = [r["response_time_ms"] for r in results]
    avg_rt = round(mean(response_times), 2)
    med_rt = round(median(response_times), 2)
    max_rt = max(response_times)
    min_rt = min(response_times)
    stddev_rt = round(stdev(response_times), 2) if len(response_times) > 1 else 0
    under_500ms = len([r for r in results if r["response_time_ms"] < 500])
    under_1s = len([r for r in results if r["response_time_ms"] < 1000])

    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total Requests", len(results)])
    ws_summary.append(["Unique Products Viewed", len(set(r['product_id'] for r in results))])
    ws_summary.append(["Successful (200)", status_codes.get(200, 0)])
    ws_summary.append(["Avg Response Time (ms)", avg_rt])
    ws_summary.append(["Median Response Time (ms)", med_rt])
    ws_summary.append(["Min Response Time (ms)", min_rt])
    ws_summary.append(["Max Response Time (ms)", max_rt])
    ws_summary.append(["Std Dev Response Time (ms)", stddev_rt])
    ws_summary.append(["< 500ms", f"{under_500ms} ({round(100 * under_500ms / len(results), 1)}%)"])
    ws_summary.append(["< 1s", f"{under_1s} ({round(100 * under_1s / len(results), 1)}%)"])

    # Performance Distribution
    ws_perf.append(["Request #", "Response Time (ms)"])
    for i, rt in enumerate(response_times, 1):
        ws_perf.append([i, rt])

    chart2 = LineChart()
    chart2.title = "Response Time Trend"
    chart2.y_axis.title = "Time (ms)"
    chart2.x_axis.title = "Request #"
    data = Reference(ws_perf, min_col=2, min_row=1, max_row=len(response_times) + 1)
    cats = Reference(ws_perf, min_col=1, min_row=2, max_row=len(response_times) + 1)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    ws_perf.add_chart(chart2, "D2")

    # Save File
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"{filename_prefix}_{timestamp}.xlsx"
    output_path = Path(folder) / filename
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"\n📊 Results exported to {output_path.absolute()}")
    return output_path.name
